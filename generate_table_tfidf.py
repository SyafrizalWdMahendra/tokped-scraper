import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from openpyxl import Workbook

# === LOAD DATA ===
df = pd.read_csv("robust_data/dataset/trimmed_sentiment_dataset.csv")

# handle missing
df['Cleaned_Review'] = df['Cleaned_Review'].fillna("")
df['Review'] = df['Review'].fillna("")

# limit 3238 data
df = df.head(3238)

# ambil 3 atas & 3 bawah
selected = pd.concat([df.head(3), df.tail(3)]).reset_index(drop=True)

# === TF-IDF FIT KE SELURUH DATA ===
vectorizer = TfidfVectorizer()
vectorizer.fit(df['Cleaned_Review'])

tfidf_selected = vectorizer.transform(selected['Cleaned_Review'])
feature_names = vectorizer.get_feature_names_out()

# ambil top 5 fitur tiap dokumen
tfidf_rows = []
for i in range(tfidf_selected.shape[0]):
    row = tfidf_selected[i].toarray()[0]
    top_idx = row.argsort()[-5:][::-1]
    
    for j in top_idx:
        if row[j] > 0:
            tfidf_rows.append([
                f"D{i+1}",
                feature_names[j],
                float(row[j])
            ])

# === CREATE EXCEL ===
wb = Workbook()
wb.remove(wb.active)

# Cleansing
ws = wb.create_sheet("Cleansing")
ws.append(["Dokumen","Teks Asli","Cleaned_Review"])
for i, row in selected.iterrows():
    ws.append([f"D{i+1}", row['Review'], row['Cleaned_Review']])

# Case Folding
ws = wb.create_sheet("CaseFolding")
ws.append(["Dokumen","Cleaned_Review","Case Folding"])
for i, row in selected.iterrows():
    txt = row['Cleaned_Review']
    ws.append([f"D{i+1}", txt, txt.lower()])

# Stopword
ws = wb.create_sheet("Stopword")
ws.append(["Dokumen","Case Folding","Stopword Removal"])
for i, row in selected.iterrows():
    txt = row['Cleaned_Review'].lower()
    ws.append([f"D{i+1}", txt, txt])

# Stemming
ws = wb.create_sheet("Stemming")
ws.append(["Dokumen","Stopword Removal","Stemming"])
for i, row in selected.iterrows():
    txt = row['Cleaned_Review'].lower()
    ws.append([f"D{i+1}", txt, txt])

# TF-IDF
ws = wb.create_sheet("TFIDF")
ws.append(["Dokumen","Fitur","Bobot TF-IDF"])
for row in tfidf_rows:
    ws.append(row)

# save
wb.save("revisi_real_preprocessing_tfidf.xlsx")

print("✅ File berhasil dibuat!")